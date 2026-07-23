"""事業性評価データを単一シートのExcelブックへ書き込む。"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

NAVY = "16324F"
BLUE = "0B5CAB"
LIGHT_BLUE = "EAF3F8"
YELLOW = "FFD43B"
LIGHT_YELLOW = "FFF8D6"
WHITE = "FFFFFF"
GRID = "B8C8D6"
MAX_COLUMNS = 7


def _title(ws: Worksheet, text: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=MAX_COLUMNS)
    cell = ws.cell(1, 1, text)
    cell.font = Font(color=WHITE, bold=True, size=16)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


def _section_header(ws: Worksheet, row_number: int, text: str) -> None:
    ws.merge_cells(
        start_row=row_number,
        start_column=1,
        end_row=row_number,
        end_column=MAX_COLUMNS,
    )
    cell = ws.cell(row_number, 1, text)
    cell.font = Font(color=WHITE, bold=True, size=12)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row_number].height = 24


def _write_table(
    ws: Worksheet,
    header_row: int,
    headers: list[str],
    rows: list[list[Any]],
    *,
    status_column: int | None = None,
) -> int:
    """表を書き込み、次の空き行番号を返す。"""
    thin = Side(style="thin", color=GRID)
    for column, value in enumerate(headers, 1):
        cell = ws.cell(header_row, column, value)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_number, values in enumerate(rows, header_row + 1):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_number, column, value)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if row_number % 2 == 0 else WHITE)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if status_column is not None:
            status_cell = ws.cell(row_number, status_column)
            status_cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

    return header_row + len(rows) + 1


def _summary_text(data: dict[str, Any], category: str) -> str:
    if category == "顧客属性":
        customer = data["customer_attributes"]
        return f"{customer.get('企業名', '')} / {customer.get('業種', '')}"
    if category == "財務分析":
        rows = data["financial_analysis"]
        return " / ".join(str(row.get("AI評価", "")) for row in rows[:2])
    if category == "取引先":
        return f"主要な販売先・仕入先 {len(data['trading_partners'])}先を確認"
    if category == "SWOT分析":
        swot = data["swot_analysis"]
        return f"強み: {swot.get('強み', {}).get('内容', '')} / 脅威: {swot.get('脅威', {}).get('内容', '')}"
    high_priority = [
        row.get("事業課題", "")
        for row in data["business_challenges"]
        if row.get("優先度") == "高"
    ]
    return " / ".join(high_priority)


def _configure_sheet(ws: Worksheet) -> None:
    widths = [18, 30, 20, 24, 26, 50, 14]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:2"


def build_business_evaluation_excel(
    data: dict[str, Any],
    source_documents: list[str],
) -> bytes:
    """確認前の事業性評価データを単一シートのExcelブックにする。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事業性評価シート"
    _title(sheet, "事業性評価シート")

    metadata = [
        ("作成状態", "要職員確認"),
        ("読取エンジン", "Azure AI Document Intelligence"),
        ("整形・分析エンジン", "Azure OpenAI"),
        ("対象企業", data["customer_attributes"].get("企業名", "")),
        ("評価基準日", data["customer_attributes"].get("評価基準日", "")),
    ]
    for row_number, (label, value) in enumerate(metadata, 3):
        sheet.cell(row_number, 1, label).font = Font(bold=True, color=NAVY)
        sheet.cell(row_number, 2, value)
    sheet["B3"].fill = PatternFill("solid", fgColor=YELLOW)
    sheet["B3"].font = Font(bold=True, color=NAVY)

    current_row = 9
    _section_header(sheet, current_row, "入力資料")
    source_rows = [[index, filename] for index, filename in enumerate(source_documents, 1)]
    current_row = _write_table(sheet, current_row + 1, ["No.", "ファイル名"], source_rows)

    current_row += 1
    _section_header(sheet, current_row, "評価概要")
    categories = ["顧客属性", "財務分析", "取引先", "SWOT分析", "事業課題"]
    summary_rows = [[category, _summary_text(data, category), "要確認"] for category in categories]
    current_row = _write_table(
        sheet,
        current_row + 1,
        ["評価領域", "AI整理結果の概要", "職員確認"],
        summary_rows,
        status_column=3,
    )

    current_row += 1
    _section_header(sheet, current_row, "顧客属性")
    customer_rows = [
        [key, value, "要確認", ""]
        for key, value in data["customer_attributes"].items()
    ]
    current_row = _write_table(
        sheet,
        current_row + 1,
        ["項目", "内容", "職員確認", "備考"],
        customer_rows,
        status_column=3,
    )

    current_row += 1
    _section_header(sheet, current_row, "財務分析")
    financial_rows = [
        [
            row.get("項目", ""),
            row.get("2024年3月期", ""),
            row.get("2025年3月期", ""),
            row.get("2026年3月期", ""),
            row.get("単位", ""),
            row.get("AI評価", ""),
            "要確認",
        ]
        for row in data["financial_analysis"]
    ]
    current_row = _write_table(
        sheet,
        current_row + 1,
        ["項目", "2024年3月期", "2025年3月期", "2026年3月期", "単位", "AI評価", "職員確認"],
        financial_rows,
        status_column=7,
    )

    current_row += 1
    _section_header(sheet, current_row, "主要な販売先・仕入先")
    partner_rows = [
        [
            row.get("区分", ""),
            row.get("取引先名", ""),
            row.get("取引比率", ""),
            row.get("決済条件", ""),
            row.get("取引状況・リスク", ""),
            "要確認",
        ]
        for row in data["trading_partners"]
    ]
    current_row = _write_table(
        sheet,
        current_row + 1,
        ["区分", "取引先名", "取引比率", "決済条件", "取引状況・リスク", "職員確認"],
        partner_rows,
        status_column=6,
    )

    current_row += 1
    _section_header(sheet, current_row, "SWOT分析")
    swot_rows = [
        [category, values.get("内容", ""), values.get("評価根拠", ""), "要確認"]
        for category, values in data["swot_analysis"].items()
    ]
    current_row = _write_table(
        sheet,
        current_row + 1,
        ["区分", "内容", "評価根拠", "職員確認"],
        swot_rows,
        status_column=4,
    )

    current_row += 1
    _section_header(sheet, current_row, "事業課題と対応方針")
    challenge_rows = [
        [
            row.get("優先度", ""),
            row.get("分類", ""),
            row.get("事業課題", ""),
            row.get("現状・背景", ""),
            row.get("対応案", ""),
            row.get("KPI", ""),
            "要確認",
        ]
        for row in data["business_challenges"]
    ]
    current_row = _write_table(
        sheet,
        current_row + 1,
        ["優先度", "分類", "事業課題", "現状・背景", "対応案", "KPI", "職員確認"],
        challenge_rows,
        status_column=7,
    )

    current_row += 1
    sheet.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=MAX_COLUMNS,
    )
    warning = sheet.cell(
        current_row,
        1,
        "注意: AIの抽出・分析結果は参考情報です。原資料と照合し、担当職員が確認・修正してから利用してください。",
    )
    warning.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    warning.font = Font(bold=True, color=NAVY)
    warning.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[current_row].height = 34

    _configure_sheet(sheet)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
