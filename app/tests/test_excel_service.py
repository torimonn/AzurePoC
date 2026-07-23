from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from services.excel_service import build_business_evaluation_excel
from services.mock_data import (
    BUSINESS_CHALLENGES,
    CUSTOMER_ATTRIBUTES,
    FINANCIAL_ANALYSIS,
    SOURCE_DOCUMENTS,
    SWOT_ANALYSIS,
    TRADING_PARTNERS,
)


class BusinessEvaluationExcelTest(unittest.TestCase):
    def setUp(self) -> None:
        self.content = build_business_evaluation_excel(
            {
                "customer_attributes": CUSTOMER_ATTRIBUTES,
                "financial_analysis": FINANCIAL_ANALYSIS,
                "trading_partners": TRADING_PARTNERS,
                "swot_analysis": SWOT_ANALYSIS,
                "business_challenges": BUSINESS_CHALLENGES,
            },
            SOURCE_DOCUMENTS,
        )
        self.workbook = load_workbook(BytesIO(self.content), read_only=True)
        self.sheet = self.workbook["事業性評価シート"]

    def _find_row(self, value: str) -> int:
        for row in self.sheet.iter_rows():
            if row[0].value == value:
                return row[0].row
        self.fail(f"{value!r} was not found in column A")

    def _find_rows(self, value: str) -> list[int]:
        return [
            row[0].row
            for row in self.sheet.iter_rows()
            if row[0].value == value
        ]

    def _find_row_after(self, value: str, after_row: int) -> int:
        for row in self.sheet.iter_rows(min_row=after_row + 1):
            if row[0].value == value:
                return row[0].row
        self.fail(f"{value!r} was not found after row {after_row}")

    def test_builds_single_business_evaluation_sheet(self) -> None:
        self.assertEqual(self.workbook.sheetnames, ["事業性評価シート"])

    def test_records_processing_pipeline_and_source_documents(self) -> None:
        self.assertEqual(self.sheet["B3"].value, "要職員確認")
        self.assertEqual(self.sheet["B4"].value, "Azure AI Document Intelligence")
        self.assertEqual(self.sheet["B5"].value, "Azure OpenAI")
        self.assertEqual(self.sheet["A11"].value, 1)
        self.assertEqual(self.sheet["B11"].value, SOURCE_DOCUMENTS[0])

    def test_contains_all_business_evaluation_sections(self) -> None:
        first_column = [row[0].value for row in self.sheet.iter_rows()]

        for section in [
            "入力資料",
            "評価概要",
            "顧客属性",
            "財務分析",
            "主要な販売先・仕入先",
            "SWOT分析",
            "事業課題と対応方針",
        ]:
            self.assertIn(section, first_column)

    def test_marks_each_detail_section_as_requiring_staff_review(self) -> None:
        customer_row = self._find_row_after("評価基準日", self._find_rows("顧客属性")[-1])
        financial_row = self._find_row_after("売上高", self._find_rows("財務分析")[-1])
        partner_row = self._find_row_after("販売先", self._find_row("主要な販売先・仕入先"))
        swot_row = self._find_row_after("強み", self._find_rows("SWOT分析")[-1])
        challenge_row = self._find_row_after("高", self._find_row("事業課題と対応方針"))

        self.assertEqual(self.sheet.cell(customer_row, 3).value, "要確認")
        self.assertEqual(self.sheet.cell(financial_row, 7).value, "要確認")
        self.assertEqual(self.sheet.cell(partner_row, 6).value, "要確認")
        self.assertEqual(self.sheet.cell(swot_row, 4).value, "要確認")
        self.assertEqual(self.sheet.cell(challenge_row, 7).value, "要確認")

    def test_writes_customer_and_financial_values(self) -> None:
        customer_row = self._find_row_after("企業名", self._find_rows("顧客属性")[-1])
        financial_row = self._find_row_after("売上高", self._find_rows("財務分析")[-1])

        self.assertEqual(self.sheet.cell(customer_row, 2).value, "株式会社 遠州テクノ")
        self.assertEqual(self.sheet.cell(financial_row, 4).value, "548,000")


if __name__ == "__main__":
    unittest.main()
